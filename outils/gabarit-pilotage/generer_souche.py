#!/usr/bin/env python3
"""Générateur de la souche `gabarit-pilotage-mission.xlsx` (chantier T-0033, temps 1/2).

Le SCRIPT est le livrable reproductible — jamais un binaire édité à la main (leçon de
gouvernance S35, docs/epreuves/2026-07-14-t-0031-consolidation-s35.md). Il régénère de zéro
la souche canon `contrats/socle/gabarit-pilotage-mission.xlsx` :

  - trois tables nommées T_Affectations / T_Imputations / T_Echeancier conformes à
    `contrats/socle/modele-donnees.md` §5.2 (noms de tables et en-têtes FIGÉS, repris
    caractère pour caractère, accents compris) ;
  - EN-TÊTES SEULS : la plage (ref) de chaque table couvre exactement la ligne d'en-tête,
    ZÉRO ligne de corps, aucune ligne vide résiduelle. C'est le remède au STOP n°3 de S35
    (« vider une ligne de tableau à la main laisse une ligne de corps VIDE » → Graph
    count:1 ≠ 0). Une table en-tête seul a une collection `rows` vide → count:0 (Graph
    workbookTable : rows = lignes de données hors en-tête) ;
  - statuts de référence ACCENTUÉS (« à valider » / « validé » ; « à émettre » / « émise »
    / « payée ») portés par des listes de validation sur les colonnes de statut — aucun
    besoin d'une ligne d'exemple pour les héberger ;
  - AUCUNE donnée d'illustration : zéro MIS-EXEMPLE, zéro jane.doe, zéro F-EXEMPLE
    (c'est la pollution que ce chantier assainit, STOP n°2 de S35).

Mises en forme reprises de la souche existante (non polluées) : en-têtes blanc Arial gras
sur fond bleu foncé, largeurs de colonnes, style de table TableStyleLight1 sans bandes.

Usage :
    python3 generer_souche.py [--sortie CHEMIN]   # génère puis auto-vérifie
    python3 generer_souche.py --verifier CHEMIN   # vérifie seulement un fichier existant

Dépendance : openpyxl (pip install openpyxl --break-system-packages).
"""
from __future__ import annotations

import argparse
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Schéma NORMATIF — modele-donnees.md §5.2 (fait foi ; en-têtes caractère pour caractère).
# ---------------------------------------------------------------------------
TABLES = [
    {
        "feuille": "Affectations",
        "table": "T_Affectations",
        "colonnes": ["CodeMission", "Ressource", "Mois", "JoursPrevus"],
        "largeurs": [17, 15, 14, 17],
    },
    {
        "feuille": "Imputations",
        "table": "T_Imputations",
        "colonnes": ["CodeMission", "Ressource", "Mois", "JoursRealises", "StatutValidation"],
        "largeurs": [17, 15, 14, 19, 22],
    },
    {
        "feuille": "Echeancier",
        "table": "T_Echeancier",
        "colonnes": ["NumFacture", "CodeMission", "MoisCA", "MontantHT", "Echeance", "Statut", "LienFacture"],
        "largeurs": [16, 17, 14, 15, 14, 14, 17],
    },
]

# Statuts de référence ACCENTUÉS — modele-donnees.md §5.2 (colonne Notes).
STATUTS_VALIDATION = ["à valider", "validé"]          # T_Imputations.StatutValidation
STATUTS_ECHEANCIER = ["à émettre", "émise", "payée"]  # T_Echeancier.Statut

# Colonnes portant une liste de validation de statut (nom de colonne -> valeurs).
VALIDATIONS_STATUT = {
    "StatutValidation": STATUTS_VALIDATION,
    "Statut": STATUTS_ECHEANCIER,
}

# Formats de nombre par en-tête (colonnes vides — style de colonne, ne crée aucune cellule).
FORMATS_COLONNE = {
    "Mois": "yyyy-mm-dd",
    "MoisCA": "yyyy-mm-dd",
    "Echeance": "yyyy-mm-dd",
    "MontantHT": "#,##0",
}

# Mises en forme (reprises de la souche existante, non polluées).
COULEUR_ENTETE = "1F3864"   # bleu foncé
POLICE = "Arial"
FILL_ENTETE = PatternFill(fill_type="solid", fgColor=COULEUR_ENTETE)
FONT_ENTETE = Font(name=POLICE, size=11, bold=True, color="FFFFFF")
ALIGN_ENTETE = Alignment(horizontal="left")

DEFAUT_SORTIE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "contrats", "socle", "gabarit-pilotage-mission.xlsx",
)


def construire_legende(ws) -> None:
    """Feuille d'aide, régénérée pour rester conforme au canon (§5.2 / §5.6).

    Corrige deux scories de l'ancienne souche : le nom d'instanciation
    (`gabarit-<CodeMission>.xlsx`, pas `pilotage-...`) et la mention d'un « consolidé
    central » caduc (le modèle v2 n'a PAS de classeur consolidé ; les coûts vivent dans le
    référentiel de coûts à audience restreinte). Ne décrit AUCUNE ligne d'exemple : les
    tables sont livrées vides (en-têtes seuls).
    """
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    lignes = [
        ("Gabarit de pilotage de mission", Font(name=POLICE, size=14, bold=True)),
        ("", None),
        ("Souche canon (modele-donnees.md §5.2). Instanciée par l'agent, une par mission, "
         "en gabarit-<CodeMission>.xlsx (instanciation SYSTÉMATIQUE et machine, §5.6).", None),
        ("Trois onglets. Les noms de tables et les en-têtes de colonnes sont FIGÉS (§5.2) : "
         "ne jamais les renommer.", None),
        ("Les tables sont livrées VIDES (en-têtes seuls). Saisir les données réelles sous "
         "la ligne d'en-tête ; aucune ligne d'exemple à remplacer.", None),
        ("Aucun coût, taux ni salaire ne figure dans ce fichier : les coûts vivent dans le "
         "référentiel de coûts à audience restreinte (§5.3), jamais dans les gabarits mission.", None),
        ("", None),
        ("T_Affectations", Font(name=POLICE, bold=True)),
        ("prévoir le staffing : qui, quel mois, combien de jours prévus.", None),
        ("T_Imputations", Font(name=POLICE, bold=True)),
        ("déclarer le réalisé : jours réalisés par ressource et par mois + statut de "
         "validation (« à valider » / « validé »).", None),
        ("T_Echeancier", Font(name=POLICE, bold=True)),
        ("plan de facturation : facture, mois de CA, montant HT, échéance, statut "
         "(« à émettre » / « émise » / « payée »), lien PDF.", None),
    ]
    for i, (texte, font) in enumerate(lignes, start=1):
        cell = ws.cell(row=i, column=1, value=texte if texte else None)
        if font is not None:
            cell.font = font


def construire_table(ws, spec: dict) -> None:
    """Écrit une feuille de table : en-têtes seuls + table nommée en-tête seul + validations."""
    colonnes = spec["colonnes"]
    largeurs = spec["largeurs"]
    n = len(colonnes)

    # Ligne d'en-tête (ligne 1) — la seule ligne écrite.
    for j, (nom, largeur) in enumerate(zip(colonnes, largeurs), start=1):
        lettre = get_column_letter(j)
        cell = ws.cell(row=1, column=j, value=nom)
        cell.font = FONT_ENTETE
        cell.fill = FILL_ENTETE
        cell.alignment = ALIGN_ENTETE
        cell.number_format = "General"  # en-tête = texte, jamais un format hérité
        ws.column_dimensions[lettre].width = largeur
        # Format par défaut de la colonne (style de colonne : ne crée AUCUNE cellule de corps).
        fmt = FORMATS_COLONNE.get(nom)
        if fmt:
            ws.column_dimensions[lettre].number_format = fmt

    # Table nommée EN-TÊTE SEUL : ref = ligne 1 uniquement -> zéro ligne de corps -> count:0.
    ref = f"A1:{get_column_letter(n)}1"
    table = Table(displayName=spec["table"], ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=False,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(table)

    # Listes de validation des statuts accentués (portées sous l'en-tête ; ne crée aucune
    # ligne de corps de table — la validation est indépendante de la plage de la table).
    for j, nom in enumerate(colonnes, start=1):
        valeurs = VALIDATIONS_STATUT.get(nom)
        if not valeurs:
            continue
        lettre = get_column_letter(j)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(valeurs) + '"',
            allow_blank=True,
            showDropDown=False,  # False = la flèche déroulante EST affichée (sémantique OOXML inversée)
        )
        dv.add(f"{lettre}2:{lettre}1048576")
        ws.add_data_validation(dv)


def generer(chemin: str) -> None:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Legende"
    construire_legende(ws0)
    for spec in TABLES:
        ws = wb.create_sheet(title=spec["feuille"])
        construire_table(ws, spec)
    wb.save(chemin)


# ---------------------------------------------------------------------------
# Auto-vérification — rouvre le fichier produit et affirme les invariants (a)..(e).
# ---------------------------------------------------------------------------
class EchecVerification(Exception):
    pass


def _affirmer(condition: bool, message: str) -> None:
    if not condition:
        raise EchecVerification(message)


def verifier(chemin: str) -> None:
    wb = load_workbook(chemin)
    print(f"Vérification de : {chemin}")

    # Index des tables par nom, toutes feuilles confondues.
    tables_par_nom = {}
    for ws in wb.worksheets:
        for nom in ws.tables:
            tables_par_nom[nom] = (ws, ws.tables[nom])

    for spec in TABLES:
        nom = spec["table"]
        colonnes = spec["colonnes"]

        # (a) la table existe sous son nom exact.
        _affirmer(nom in tables_par_nom, f"(a) table absente : {nom}")
        ws, table = tables_par_nom[nom]
        print(f"  (a) table {nom!r} présente sur la feuille {ws.title!r} — OK")

        # (b) ref = ligne d'en-tête SEULE (min_row == max_row == 1).
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        _affirmer(
            min_row == 1 and max_row == 1,
            f"(b) {nom} : ref {table.ref} n'est pas en-tête seul (lignes {min_row}..{max_row})",
        )
        print(f"  (b) {nom} ref={table.ref} = en-tête seul (1 ligne) — OK")

        # (c) aucune cellule non vide sous l'en-tête (ligne >= 2) sur la feuille.
        residus = [
            c.coordinate
            for row in ws.iter_rows(min_row=2)
            for c in row
            if c.value is not None
        ]
        _affirmer(not residus, f"(c) {nom} : cellules non vides sous l'en-tête : {residus}")
        print(f"  (c) {nom} : aucune cellule sous l'en-tête (max_row={ws.max_row}) — OK")

        # (d) en-têtes caractère pour caractère == §5.2 (colonnes de table ET cellules).
        noms_colonnes = [c.name for c in table.tableColumns]
        _affirmer(
            noms_colonnes == colonnes,
            f"(d) {nom} : colonnes de table {noms_colonnes} != §5.2 {colonnes}",
        )
        cellules = [ws.cell(row=1, column=j).value for j in range(min_col, max_col + 1)]
        _affirmer(
            cellules == colonnes,
            f"(d) {nom} : cellules d'en-tête {cellules} != §5.2 {colonnes}",
        )
        print(f"  (d) {nom} en-têtes == §5.2 : {colonnes} — OK")

    # (e) statuts de référence ACCENTUÉS présents dans les listes de validation.
    formules = []
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            if dv.formula1:
                formules.append(dv.formula1)
    blob = " || ".join(formules)
    attendus = STATUTS_VALIDATION + STATUTS_ECHEANCIER
    manquants = [s for s in attendus if s not in blob]
    _affirmer(not manquants, f"(e) statuts accentués absents des validations : {manquants}")
    print(f"  (e) statuts accentués présents dans les validations : {attendus} — OK")

    # Garde-fou anti-régression S35 : aucune trace des données d'illustration polluantes.
    interdits = ["MIS-EXEMPLE", "jane.doe", "F-EXEMPLE", "a valider", "a emettre", "a emmettre"]
    trouves = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str):
                    for mot in interdits:
                        if mot in c.value:
                            trouves.append(f"{ws.title}!{c.coordinate}={c.value!r}")
    _affirmer(not trouves, f"(pollution) données/statuts non conformes trouvés : {trouves}")
    print("  (anti-pollution) aucune donnée d'illustration ni statut non accentué — OK")

    print("VÉRIFICATION RÉUSSIE — souche conforme §5.2 (en-têtes seuls, statuts accentués).")


def main() -> int:
    parser = argparse.ArgumentParser(description="Génère/vérifie la souche gabarit-pilotage-mission.xlsx.")
    parser.add_argument("--sortie", default=DEFAUT_SORTIE, help="Chemin du fichier à générer.")
    parser.add_argument("--verifier", metavar="CHEMIN", help="Vérifier seulement un fichier existant (pas de génération).")
    args = parser.parse_args()

    try:
        if args.verifier:
            verifier(args.verifier)
            return 0
        generer(args.sortie)
        print(f"Souche générée : {args.sortie}")
        verifier(args.sortie)  # auto-vérification systématique en fin de génération
        return 0
    except EchecVerification as e:
        print(f"ÉCHEC DE VÉRIFICATION : {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
